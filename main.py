# -*- coding: utf-8 -*-
"""
Created on Sun Aug 17 19:40:11 2025

@author: Hugo
"""

# app.py
# -*- coding: utf-8 -*-
from __future__ import annotations
import os, re, json, base64, traceback
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urlparse, unquote
from typing import Any, List, Tuple, Optional, Union, Dict
from datetime import datetime
from copy import copy

import requests
from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware

import firebase_admin
from firebase_admin import credentials, firestore
from google.cloud import storage as gcs

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Border, Side

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from PIL import Image as PILImage

# ───────────────────────────────── CONFIG ──────────────────────────────────
# Template en el root del repo (override con RENDIDOR_TEMPLATE_PATH)
TEMPLATE_XLSX_DEFAULT = Path(
    os.getenv("RENDIDOR_TEMPLATE_PATH", Path(__file__).parent / "FormatoRendicion.xlsx")
).resolve()

# Salida a /tmp por defecto (filesystem efímero de Render)
OUTPUT_DIR = Path(os.getenv("RENDIDOR_OUTPUT_DIR", "/tmp/rendidor")).resolve()
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

def default_fotos_dir() -> Path:
    p = OUTPUT_DIR / "fotos_rendicion"
    p.mkdir(parents=True, exist_ok=True)
    return p

DEFAULT_NAME_CAMPANA = "—"
DEFAULT_NAME_PERSONA = "—"

MESES_ES = [
    "enero","febrero","marzo","abril","mayo","junio",
    "julio","agosto","septiembre","octubre","noviembre","diciembre"
]

# ─────────────────────────────── FastAPI init ───────────────────────────────
app = FastAPI(title="Rendidor API", version="1.0.0")

# Middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], allow_methods=["GET"], allow_headers=["*"]
)

# ───────────────────────── Firestore / Storage init ─────────────────────────
def _get_sa_info() -> dict:
    b64 = os.getenv("FIREBASE_KEY_B64", "").strip()
    if not b64:
        raise RuntimeError(
            "FIREBASE_KEY_B64 no está definida. Sube tu service-account.json como base64 en esta variable."
        )
    try:
        return json.loads(base64.b64decode(b64))
    except Exception as e:
        raise RuntimeError(f"FIREBASE_KEY_B64 inválida: {e}")

def init_clients() -> Tuple[firestore.Client, gcs.Client]:
    sa = _get_sa_info()
    cred = credentials.Certificate(sa)
    if not firebase_admin._apps:
        firebase_admin.initialize_app(cred)
    fs_client = firestore.client()
    gcs_client = gcs.Client.from_service_account_info(sa)
    return fs_client, gcs_client

def _unique_path(path: Path) -> Path:
    path = Path(path)
    if not path.exists():
        return path
    stem, suffix = path.stem, path.suffix
    i = 1
    while True:
        candidate = path.with_name(f"{stem} ({i}){suffix}")
        if not candidate.exists():
            return candidate
        i += 1

# ───────────────────────────── Helpers comunes ──────────────────────────────
def _extract_urls_from_fotos(fotos: Any) -> List[str]:
    urls: List[str] = []
    if not fotos: return urls
    def _add(x: Any):
        if not x: return
        if isinstance(x, str):
            urls.append(x.strip())
        elif isinstance(x, dict):
            for k in ("url", "href", "downloadURL", "path"):
                if k in x and isinstance(x[k], str):
                    urls.append(x[k].strip())
    if isinstance(fotos, list):
        for item in fotos: _add(item)
    else:
        _add(fotos)
    return list(dict.fromkeys(urls))

def _safe_ext_from_url(u: str, default=".jpg") -> str:
    try:
        name = unquote(Path(urlparse(u).path).name)
        if "." in name:
            return "." + name.split(".")[-1].split("?")[0][:8]
    except Exception:
        pass
    return default

def _safe_filename(base: str, idx: int, url: str) -> str:
    ext = _safe_ext_from_url(url)
    base = re.sub(r"[^a-zA-Z0-9_\-\.]+", "_", base)[:64]
    return f"{base}_{idx:03d}{ext}"

def _download_http(url: str, dest: Path) -> None:
    with requests.get(url, stream=True, timeout=30) as r:
        r.raise_for_status()
        dest.parent.mkdir(parents=True, exist_ok=True)
        with open(dest, "wb") as f:
            for chunk in r.iter_content(chunk_size=8192):
                if chunk: f.write(chunk)

def _download_gs(gcs_client: gcs.Client, gs_url: str, dest: Path) -> None:
    path = gs_url[5:]
    bucket_name, _, blob_path = path.partition("/")
    bucket = gcs_client.bucket(bucket_name)
    blob = bucket.blob(blob_path)
    dest.parent.mkdir(parents=True, exist_ok=True)
    blob.download_to_filename(str(dest))

def _is_gs(u: str) -> bool:
    return u.startswith("gs://")

def _parse_us_dt(x: Any) -> Optional[datetime]:
    if x is None: return None
    if isinstance(x, datetime): return x
    if isinstance(x, str):
        s = x.strip()
        for fmt in ("%m-%d-%Y %H:%M:%S", "%m-%d-%Y %H:%M", "%m-%d-%Y"):
            try:
                return datetime.strptime(s, fmt)
            except Exception:
                continue
    return None

def _to_ddmmyyyy(dt: datetime) -> str:
    return dt.strftime("%d-%m-%Y")

def _clean_monto(x: Any) -> float:
    if x is None: return 0.0
    if isinstance(x, (int, float)): return float(x)
    s = str(x).strip()
    s = re.sub(r"[^\d,.\-]", "", s)
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0

def mes_es(dt: datetime) -> str:
    return MESES_ES[dt.month - 1]

# ───────────── Firestore lectura / normalización ─────────────
def fetch_registros(
    collection_name: str = "registros",
    fotos_field: str = "foto",
    campana_id: Optional[str] = None,
) -> Tuple[List[dict], Dict[str, Any]]:
    fs, _ = init_clients()
    registros: List[dict] = []
    nameCampana = None; namePersona = None; startDC = None; endDC = None
    montoTotal = 0.0

    q = fs.collection(collection_name)
    if campana_id:
        q = q.where("campanaID", "==", campana_id)

    for doc in q.stream():
        data = doc.to_dict() or {}
        d_dt = _parse_us_dt(data.get("date")) or _parse_us_dt(data.get("Fecha"))
        cat = data.get("categoria") or data.get("Detalle") or ""
        idb = data.get("idBoleta") or data.get("Boleta") or ""
        mon = _clean_monto(data.get("monto") or data.get("Gasto"))
        fotos = _extract_urls_from_fotos(data.get(fotos_field))
        registros.append({
            "doc_id": doc.id, "date": d_dt, "categoria": cat,
            "idBoleta": idb, "Monto": mon, "fotos": fotos
        })
        if nameCampana is None: nameCampana = data.get("nameCampana")
        if namePersona is None: namePersona = data.get("namePersona")
        if startDC is None:    startDC = _parse_us_dt(data.get("startDateCampana"))
        if endDC   is None:    endDC   = _parse_us_dt(data.get("endDateCampana"))
        if not montoTotal:     montoTotal = _clean_monto(data.get("montoTotal"))

    fechas_validas = [r["date"] for r in registros if r["date"] is not None]
    if startDC is None and fechas_validas: startDC = min(fechas_validas)
    if endDC   is None and fechas_validas: endDC   = max(fechas_validas)

    meta = {
        "nameCampana": nameCampana or DEFAULT_NAME_CAMPANA,
        "namePersona": namePersona or DEFAULT_NAME_PERSONA,
        "startDateCampana": startDC,
        "endDateCampana": endDC,
        "montoTotal": montoTotal,
    }
    registros.sort(key=lambda r: (r["date"] or datetime(1900,1,1), r["doc_id"]))
    return registros, meta

# ─────── Excel: headers / estilo / escritura desde plantilla ───────
def _find_headers(ws: Worksheet, names=("Fecha","Detalle","Boleta","Gasto"), search_rows: int = 20) -> Dict[str,int]:
    headers: Dict[str,int] = {}; wanted = set(names)
    for r in range(1, search_rows + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str):
                v = val.strip()
                if v in wanted and v not in headers:
                    headers[v] = c
        if len(headers) == len(names): break
    defaults = {"Fecha":1, "Detalle":2, "Boleta":3, "Gasto":4}
    for k in defaults: headers.setdefault(k, defaults[k])
    return headers

def _copy_row_style(ws, src_row: int, dst_row: int, cols: list[int]) -> None:
    try: ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    except Exception: pass
    for col in cols:
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        try:
            if src.has_style:
                try: dst.font = copy(src.font)
                except Exception: pass
                try: dst.fill = copy(src.fill)
                except Exception: pass
                try: dst.border = copy(src.border)
                except Exception: pass
                try: dst.alignment = copy(src.alignment)
                except Exception: pass
                try: dst.protection = copy(src.protection)
                except Exception: pass
                try: dst.number_format = src.number_format
                except Exception: pass
        except TypeError:
            try: dst.number_format = src.number_format
            except Exception: pass

def write_excel_from_template(
    template_path: Union[str, Path],
    registros: List[dict],
    meta: Dict[str, Any],
    start_row: int = 11,
    sheet_name: Optional[str] = None,
) -> Path:
    template_path = Path(template_path)
    if not template_path.exists():
        raise FileNotFoundError(f"No existe la plantilla: {template_path}")

    wb = load_workbook(template_path, data_only=False)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    cols = _find_headers(ws, names=("Fecha", "Detalle", "Boleta", "Gasto"))

    nameCampana = meta.get("nameCampana") or DEFAULT_NAME_CAMPANA
    namePersona = meta.get("namePersona") or DEFAULT_NAME_PERSONA
    startDC: Optional[datetime] = meta.get("startDateCampana")
    endDC: Optional[datetime] = meta.get("endDateCampana")
    montoTotal = meta.get("montoTotal") or 0.0

    ws["C4"] = nameCampana
    ws["C5"] = namePersona
    if startDC: ws["E4"] = _to_ddmmyyyy(startDC)
    if endDC:   ws["E5"] = _to_ddmmyyyy(endDC)
    ws["E7"] = _clean_monto(montoTotal)

    used_cols = [cols["Fecha"], cols["Detalle"], cols["Boleta"], cols["Gasto"]]
    n = len(registros)

    if n >= 2:
        ws.insert_rows(start_row + 1, amount=n - 1)

    for r in range(start_row, start_row + max(n, 1)):
        _copy_row_style(ws, start_row, r, used_cols)

    for idx, reg in enumerate(registros):
        row = start_row + idx
        dt = reg["date"]
        ws.cell(row=row, column=cols["Fecha"]).value = _to_ddmmyyyy(dt) if dt else ""
        ws.cell(row=row, column=cols["Detalle"]).value = reg.get("categoria", "")
        ws.cell(row=row, column=cols["Boleta"]).value = reg.get("idBoleta", "")
        ws.cell(row=row, column=cols["Gasto"]).value = float(reg.get("Monto") or 0.0)

    # Bordes A..G
    thin = Side(style="thin", color="000000")
    box  = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(start_row, start_row + n):
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = box

    # Totales dinámicos
    first_row = start_row
    last_row  = start_row + max(n, 1) - 1
    sum_row   = (start_row + n) if n >= 1 else (start_row + 1)
    gasto_col_idx    = cols["Gasto"]
    gasto_col_letter = get_column_letter(gasto_col_idx)
    sum_cell = f"G{sum_row}"

    try: prev_fmt = ws[sum_cell].number_format
    except Exception: prev_fmt = None

    if n >= 1:
        ws[sum_cell].value = f"=SUMA({gasto_col_letter}{first_row}:{gasto_col_letter}{last_row})"
    else:
        ws[sum_cell].value = 0
    if prev_fmt:
        try: ws[sum_cell].number_format = prev_fmt
        except Exception: pass

    # G14 = SUMA(...)
    g14_new_row = 14 + max(n - 1, 0)
    g14_cell    = f"G{g14_new_row}"
    ws[g14_cell].value = f"={sum_cell}"
    try: ws[g14_cell].number_format = ws[sum_cell].number_format
    except Exception: pass

    # G16 = E7 - G14
    g16_new_row = 16 + max(n - 1, 0)
    g16_cell    = f"G{g16_new_row}"
    ws[g16_cell].value = f"=E7-{g14_cell}"
    try: ws[g16_cell].number_format = ws[g14_cell].number_format
    except Exception: pass

    # Nombre de salida
    if startDC:
        startDay = str(int(startDC.strftime("%d"))); startMonth = startDC.strftime("%m"); startYear = startDC.strftime("%Y")
    else:
        startDay, startMonth, startYear = "01","01","1900"
    endDay = str(int(endDC.strftime("%d"))) if endDC else startDay

    out_name = f"{nameCampana} Rendición {startDay}-{endDay} {startMonth} {startYear}.xlsx"
    out_path = _unique_path(OUTPUT_DIR / out_name)
    try:
        wb.save(out_path)
    except PermissionError:
        out_path = _unique_path(out_path)
        wb.save(out_path)
    return out_path

# ─────────── Descarga fotos y armado de PDF por fecha ───────────
def download_images_grouped_by_date(
    registros: List[dict],
    gcs_client: gcs.Client,
    out_dir: Union[str, Path],
    max_workers: int = 8,
) -> Dict[str, List[Path]]:
    out_dir = Path(out_dir); out_dir.mkdir(parents=True, exist_ok=True)
    tasks = []
    for reg in registros:
        dt = reg["date"] or datetime(1900, 1, 1)
        fecha_key = _to_ddmmyyyy(dt)
        for i, u in enumerate(reg.get("fotos", [])):
            fname = _safe_filename(f"{reg['doc_id']}_{fecha_key.replace('-', '')}", i, u)
            dest = out_dir / fecha_key / fname
            tasks.append((u, dest, fecha_key))

    results: List[Tuple[str, Path, str, bool, Optional[str]]] = []
    def _one(u: str, dest: Path, fkey: str):
        try:
            if _is_gs(u): _download_gs(gcs_client, u, dest)
            else:        _download_http(u, dest)
            return (u, dest, fkey, True, None)
        except Exception as e:
            return (u, dest, fkey, False, str(e))

    if tasks:
        with ThreadPoolExecutor(max_workers=max_workers) as ex:
            futs = [ex.submit(_one, u, d, fk) for (u, d, fk) in tasks]
            for f in as_completed(futs):
                results.append(f.result())

    grouped: Dict[str, List[Path]] = {}
    for (u, p, fkey, ok, err) in results:
        if ok: grouped.setdefault(fkey, []).append(p)
    for k in grouped: grouped[k].sort()

    def _key(k: str) -> datetime: return datetime.strptime(k, "%d-%m-%Y")
    grouped = dict(sorted(grouped.items(), key=lambda kv: _key(kv[0])))
    return grouped

def build_pdf_from_grouped_images(grouped: Dict[str, List[Path]], out_pdf: Union[str, Path]) -> Path:
    out_pdf = _unique_path(Path(out_pdf))
    c = canvas.Canvas(str(out_pdf), pagesize=A4)
    pw, ph = A4
    margin = 36

    if not grouped:
        c.setFont("Helvetica-Bold", 18)
        c.drawCentredString(pw/2, ph/2, "Sin fotos")
        c.showPage()
        c.save()
        return out_pdf

    for fecha_key, paths in grouped.items():
        c.setFont("Helvetica-Bold", 20)
        c.drawCentredString(pw / 2, ph / 2, fecha_key)
        c.showPage()
        for p in paths:
            try:
                with PILImage.open(p) as im:
                    w, h = im.size
                max_w = pw - 2 * margin
                max_h = ph - 2 * margin
                scale = min(max_w / w, max_h / h)
                new_w, new_h = w * scale, h * scale
                x = (pw - new_w) / 2
                y = (ph - new_h) / 2
                c.drawImage(ImageReader(str(p)), x, y, width=new_w, height=new_h, preserveAspectRatio=True, anchor='c')
                c.showPage()
            except Exception:
                continue

    c.save()
    return out_pdf

# ───────────────────────── Lógica de orquestación ──────────────────────────
def generar_rendicion_bundle(
    campana_id: Optional[str],
    template_path: Optional[str] = None,
    collection_name: str = "registros",
    fotos_field: str = "foto",
    start_row: int = 11,
    sheet_name: Optional[str] = None,
) -> Tuple[Path, Path]:
    template = Path(template_path or TEMPLATE_XLSX_DEFAULT)
    fotos_dir = default_fotos_dir()

    registros, meta = fetch_registros(
        collection_name=collection_name,
        fotos_field=fotos_field,
        campana_id=campana_id,
    )

    excel_out = write_excel_from_template(template, registros, meta, start_row=start_row, sheet_name=sheet_name)
    _, gcs_client = init_clients()
    grouped = download_images_grouped_by_date(registros, gcs_client, out_dir=fotos_dir, max_workers=8)

    startDC = meta.get("startDateCampana")
    endDC = meta.get("endDateCampana") or startDC
    if startDC:
        startDay = str(int(startDC.strftime("%d")))
        endDay = str(int((endDC or startDC).strftime("%d")))
        mes = mes_es(startDC); yy = startDC.strftime("%y")
        pdf_name = f"{meta.get('nameCampana', '—')} Rendición {startDay}-{endDay} {mes} {yy}.pdf"
    else:
        pdf_name = "Rendicion_Fotos.pdf"

    pdf_out = build_pdf_from_grouped_images(grouped, OUTPUT_DIR / pdf_name)
    return excel_out, pdf_out

# ───────────────────────────────── Endpoints ────────────────────────────────
@app.get("/health")
def health():
    return {"status": "ok"}

@app.get("/rendicion")
def rendicion(request: Request, campana_id: str):
    try:
        excel_out, pdf_out = generar_rendicion_bundle(
            campana_id=campana_id,
            template_path=None,
            collection_name="registros",
            fotos_field="foto",
            start_row=11,
            sheet_name=None,
        )
    except FileNotFoundError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {e}")

    excel_url = str(request.url_for("download_file", kind="excel", filename=excel_out.name))
    pdf_url   = str(request.url_for("download_file", kind="pdf",   filename=pdf_out.name))
    return JSONResponse({"excel_url": excel_url, "pdf_url": pdf_url})

@app.get("/download/{kind}/{filename}")
def download_file(kind: str, filename: str):
    if kind not in {"excel", "pdf"}:
        raise HTTPException(status_code=400, detail="kind inválido")

    base_dir = OUTPUT_DIR
    file_path = (base_dir / filename).resolve()

    try:
        file_path.relative_to(base_dir)
    except Exception:
        raise HTTPException(status_code=400, detail="Ruta no permitida")

    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado")

    media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if kind=="excel" else "application/pdf"
    return FileResponse(path=str(file_path), media_type=media, filename=file_path.name)

# (en Render: START ➜ uvicorn app:app --host 0.0.0.0 --port $PORT)


